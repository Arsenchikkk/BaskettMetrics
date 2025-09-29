import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import re

# --- КОНФИГУРАЦИЯ ---
DB_HOST = "localhost"
DB_NAME = "nba db"  
DB_USER = "postgres" 
DB_PASS = "0000" 

# --- УТИЛИТАРНАЯ ФУНКЦИЯ ДЛЯ БАЗЫ ДАННЫХ ---

def run_query_to_df(query):
    """
    Выполняет SQL-запрос и ВОЗВРАЩАЕТ DataFrame.
    """
    conn = None
    try:
        conn = psycopg2.connect(host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASS)
        df = pd.read_sql(query, conn) 
        
        print(f"\n[INFO] Успешно получено {len(df)} строк.")
        return df
            
    except (Exception, psycopg2.DatabaseError) as error:
        print(f"\n[ОШИБКА DB] При выполнении запроса (проверьте БД или запрос): {error}")
        return pd.DataFrame() # Возвращаем пустой DF в случае ошибки
    finally:
        if conn:
            conn.close()

# --- ЗАДАНИЕ 1: ВИЗУАЛИЗАЦИИ (60 БАЛЛОВ) ---

def create_chart(df, title, x_label, y_label, chart_type, filename, **kwargs):
    """Строит и сохраняет график, выводит отчет в консоль."""
    if df.empty or len(df) < 2:
        print(f"[ОТЧЕТ] График '{title}' пропущен из-за недостатка данных.")
        return

    plt.figure(figsize=(12, 7))
    
    x_col = df.columns[0]
    if len(df.columns) >= 2:
        y_col = df.columns[1]
    else:
        y_col = x_col 
    
    # --- ЛОГИКА ГРАФИКОВ ---
    if chart_type == 'pie':
        plt.pie(df[y_col], labels=df[x_col], autopct='%1.1f%%', startangle=90, **kwargs)
        plt.title(f"Круговая диаграмма: {title}")
        plt.legend(title=x_label, loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    elif chart_type == 'bar':
        plt.bar(df[x_col], df[y_col], **kwargs)
        plt.title(f"Столбчатая диаграмма: {title}")
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.xticks(rotation=45, ha='right')

    elif chart_type == 'barh':
        plt.barh(df[x_col], df[y_col], **kwargs)
        plt.title(f"Горизонтальная диаграмма: {title}")
        plt.xlabel(y_label)
        plt.ylabel(x_label)

    elif chart_type == 'line':
        plt.plot(df[x_col], df[y_col], marker='o', **kwargs)
        plt.title(f"Линейный график: {title}")
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.grid(True, linestyle='--', alpha=0.7)

    elif chart_type == 'scatter':
        plt.scatter(df[x_col], df[y_col], **kwargs)
        plt.title(f"Диаграмма рассеяния: {title}")
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.grid(True, linestyle='--', alpha=0.7)
    
    # --- СОХРАНЕНИЕ И ОТЧЕТ ---
    plt.gca().spines['top'].set_visible(False)
    plt.gca().spines['right'].set_visible(False)
    plt.tight_layout()
    
    os.makedirs("charts", exist_ok=True)
    save_path = os.path.join("charts", filename)
    plt.savefig(save_path)
    plt.close()
    
    print("-" * 50)
    print(f"[ОТЧЕТ] График: {title}")
    print(f"  Тип: {chart_type}")
    print(f"  Создан файл: {save_path}")
    print(f"  Получено строк: {len(df)}")
    print(f"  Показывает: {title}. Анализ {x_label} и {y_label}.")
    print("-" * 50)

def generate_all_charts():
    print("\n\n===== ЗАДАНИЕ 1: ПОСТРОЕНИЕ ГРАФИКОВ (MATPLOTLIB) =====")

    # 1. Круговая диаграмма: Распределение событий по стартовым позициям
    query_1 = """
    SELECT
        "START_POSITION" AS position,
        COUNT(*) AS event_count
    FROM
        games_details
    WHERE
        "START_POSITION" IS NOT NULL
    GROUP BY
        1
    ORDER BY
        2 DESC;
    """
    df_1 = run_query_to_df(query_1)
    df_1 = df_1.rename(columns={'position': 'Позиция', 'event_count': 'Количество событий'})
    create_chart(df_1, 
                 "Распределение событий по стартовым позициям", 
                 "Позиция", "Количество событий", 
                 'pie', "1_start_position_pie.png",
                 textprops={'fontsize': 10} 
    )
    
    # 2. Столбчатая диаграмма: Топ-10 команд по среднему количеству передач (AST)
    query_2 = """
    SELECT
        t."NICKNAME" AS team_nickname,
        ROUND(AVG(gd."AST"), 2) AS avg_assists
    FROM
        games_details gd
    INNER JOIN
        teams t ON gd."TEAM_ID" = t."TEAM_ID"
    GROUP BY
        1
    ORDER BY
        2 DESC
    LIMIT 10;
    """
    df_2 = run_query_to_df(query_2)
    create_chart(df_2, 
                 "Топ-10 команд по среднему количеству передач (AST)", 
                 "Команда", "Среднее AST за игру", 
                 'bar', "2_top10_assists.png",
                 color='skyblue'
    )

    # 3. Горизонтальная столбчатая диаграмма: Топ-10 игроков по суммарным подборам (REB)
    query_3 = """
    SELECT
        p."PLAYER_NAME",
        COALESCE(SUM(gd."REB"), 0) AS total_rebounds
    FROM
        games_details gd
    INNER JOIN
        players p ON gd."PLAYER_ID" = p."PLAYER_ID"
    WHERE 
        gd."REB" IS NOT NULL
    GROUP BY
        1
    ORDER BY
        2 DESC
    LIMIT 10;
    """
    df_3 = run_query_to_df(query_3)
    df_3 = df_3.iloc[::-1] 
    create_chart(df_3, 
                 "Топ-10 игроков по суммарным подборам (REB)", 
                 "Игрок", "Суммарные подборы", 
                 'barh', "3_top10_rebounds.png",
                 color='salmon'
    )
    
    # 4. Линейный график: Динамика среднего процента побед по сезонам для Западной Конференции
    query_4 = """
    SELECT
        "SEASON_ID" AS season,
        AVG("W_PCT") AS avg_win_percentage
    FROM
        ranking
    WHERE
        "CONFERENCE" = 'West' 
    GROUP BY
        1
    ORDER BY
        1;
    """
    df_4 = run_query_to_df(query_4)
    df_4['season'] = df_4['season'].astype(str).str[1:].astype(int) 
    create_chart(df_4, 
                 "Динамика среднего процента побед (W_PCT) Западной Конференции по сезонам", 
                 "Сезон", "Средний % побед", 
                 'line', "4_win_pct_line.png",
                 color='green'
    )

    # 5. Сравнение 3-очковых бросков (FG3_PCT) по конференциям
    query_5 = """
    SELECT
        r."CONFERENCE" AS conference,
        ROUND(AVG(gd."FG3_PCT"), 4) AS avg_three_point_pct
    FROM
        games_details gd
    INNER JOIN
        teams t ON gd."TEAM_ID" = t."TEAM_ID"
    INNER JOIN
        ranking r ON t."TEAM_ID" = r."TEAM_ID" 
        AND r."SEASON_ID" = (SELECT MAX("SEASON_ID") FROM ranking)
    WHERE
        r."CONFERENCE" IS NOT NULL AND gd."FG3_PCT" IS NOT NULL
    GROUP BY
        1;
    """
    df_5 = run_query_to_df(query_5)
    create_chart(df_5, 
                 "Сравнение среднего % 3-очковых бросков (FG3_PCT) по конференциям", 
                 "Конференция", "Средний % 3-очковых", 
                 'bar', "5_fg3pct_conference.png",
                 color=['gold', 'darkblue']
    )

    # 6. Топ-10 лучших "Клатчеров" (Превышение среднего командного PTS)
    query_6 = """
    WITH Player_Stats AS (
        SELECT
            p."PLAYER_NAME" AS player_name, 
            t."TEAM_ID",
            ROUND(AVG(gd."PTS"), 2) AS player_avg_pts
        FROM
            games_details gd
        INNER JOIN players p ON gd."PLAYER_ID" = p."PLAYER_ID"
        INNER JOIN teams t ON gd."TEAM_ID" = t."TEAM_ID"
        WHERE 
            gd."PTS" IS NOT NULL 
            AND gd."MIN" IS NOT NULL
            -- Финальный безопасный способ обработки столбца MIN (текст -> число)
            AND CAST(CAST(COALESCE(SPLIT_PART(gd."MIN", ':', 1), '0') AS NUMERIC) AS INTEGER) > 10
        GROUP BY 1, 2
        HAVING COUNT(gd."GAME_ID") > 50
    ),
    Team_Avg AS (
        SELECT
            "TEAM_ID",
            ROUND(AVG("PTS") OVER (PARTITION BY "TEAM_ID"), 2) AS team_avg_pts
        FROM
            games_details
        GROUP BY "TEAM_ID", "PTS"
    )
    SELECT
        ps.player_name AS player, 
        ROUND(ps.player_avg_pts - ta.team_avg_pts, 2) AS pts_above_team_avg
    FROM
        Player_Stats ps
    INNER JOIN
        Team_Avg ta ON ps."TEAM_ID" = ta."TEAM_ID"
    GROUP BY
        1, 2
    ORDER BY
        2 DESC
    LIMIT 10;
    """
    df_6 = run_query_to_df(query_6)
    df_6 = df_6.iloc[::-1] # Инвертируем для barh
    create_chart(df_6, 
                 "Топ-10 лучших 'Клатчеров' (Превышение среднего PTS команды)", 
                 "Игрок", "Превышение среднего PTS команды", 
                 'barh', "6_clutch_players.png",
                 color='purple'
    )

    # 7. Диаграмма рассеяния: Взаимосвязь среднего PTS и среднего FG_PCT для всех игроков
    query_7 = """
    SELECT
        ROUND(AVG(gd."PTS"), 2) AS avg_pts,
        ROUND(AVG(gd."FG_PCT"), 4) AS avg_fg_pct
    FROM
        games_details gd
    WHERE
        gd."PTS" IS NOT NULL
        AND gd."FG_PCT" IS NOT NULL
    GROUP BY
        gd."PLAYER_ID"
    HAVING
        COUNT(gd."GAME_ID") > 100 -- Минимум 100 игр для достоверности
    ORDER BY
        1 DESC
    LIMIT 100; -- Ограничим до 100 игроков для наглядности
    """
    df_7 = run_query_to_df(query_7)
    
    # Нормализуем FG_PCT для лучшей визуализации (умножим на 100)
    df_7['avg_fg_pct'] = df_7['avg_fg_pct'] * 100 
    
    create_chart(df_7, 
                 "Взаимосвязь среднего PTS и среднего FG_PCT", 
                 "Среднее PTS за игру", "Средний % попаданий с игры (FG_PCT, %)", 
                 'scatter', "7_pts_vs_fgpct_scatter.png",
                 color='blue', alpha=0.6, s=50
    )


# --- ЗАДАНИЕ 2: ВРЕМЕННОЙ ПОЛЗУНОК (15 БАЛЛОВ) ---

# --- ЗАДАНИЕ 2: ВРЕМЕННОЙ ПОЛЗУНОК (15 БАЛЛОВ) ---

# --- ЗАДАНИЕ 2: ВРЕМЕННОЙ ПОЛЗУНОК (15 БАЛЛОВ) ---

# --- ЗАДАНИЕ 2: ВРЕМЕННОЙ ПОЛЗУНОК (15 БАЛЛОВ) ---

# --- ЗАДАНИЕ 2: ВРЕМЕННОЙ ПОЛЗУНОК (15 БАЛЛОВ) ---

def create_plotly_animation():
    print("\n\n===== ЗАДАНИЕ 2: ВРЕМЕННОЙ ПОЛЗУНОК (PLOTLY) =====")
    
    query_plotly = """
    SELECT 
        g."GAME_DATE_EST" AS game_date, 
        t."NICKNAME" AS team_name, 
        g."PTS_home" AS points
    FROM 
        games g
    INNER JOIN 
        teams t ON g."HOME_TEAM_ID" = t."TEAM_ID" 
    ORDER BY 
        g."GAME_DATE_EST"
    LIMIT 5000;
    """
    
    df = run_query_to_df(query_plotly)
    if df.empty:
        print("[ОТЧЕТ] Plotly пропущен из-за пустых данных.")
        return

    # Подготовка данных для анимации
    df['GAME_DATE_EST'] = pd.to_datetime(df['game_date'])
    df['Year'] = df['GAME_DATE_EST'].dt.year.astype(str) 

    # *** ФИНАЛЬНОЕ ИСПРАВЛЕНИЕ ***
    # УДАЛЯЕМ NaN или заменяем их на 0 в колонке 'points', которая используется для 'size'.
    df['points'] = df['points'].fillna(0)
    # *****************************

    fig = px.scatter(df, x="GAME_DATE_EST", y="points",
                     animation_frame="Year",
                     size="points", 
                     color="team_name",
                     hover_name="team_name",
                     log_x=False, 
                     size_max=20,
                     title="Динамика очков домашних команд по годам")

    fig.show()
    print("\n[ОТЧЕТ] Создан интерактивный график Plotly с временным ползунком. Проверьте в браузере.")
# --- ЗАДАНИЕ 3: ЭКСПОРТ В EXCEL С ФОРМАТИРОВАНИЕМ (25 БАЛЛОВ) ---

# --- ЗАДАНИЕ 3: ЭКСПОРТ В EXCEL С ФОРМАТИРОВАНИЕМ (25 БАЛЛОВ) ---

def export_to_excel(dataframes_dict, filename):
    """Экспортирует словарь DataFrame'ов в Excel с форматированием."""
    print("\n\n===== ЗАДАНИЕ 3: ЭКСПОРТ В EXCEL С ФОРМАТИРОВАНИЕМ =====")
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", filename)
    total_rows = 0

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            
            for sheet_name, df in dataframes_dict.items():
                if df.empty: continue
                
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                
                workbook = writer.book
                ws = workbook[sheet_name]
                total_rows += len(df)
                
                data_range_full = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
                
                ws.auto_filter.ref = data_range_full
                ws.freeze_panes = "A2" 
                
                for col_idx, dtype in enumerate(df.dtypes):
                    col_letter = get_column_letter(col_idx + 1)
                    
                    # Применяем форматирование только к числовым колонкам с более чем одной строкой данных (для градиента)
                    if pd.api.types.is_numeric_dtype(dtype) and len(df) > 1:
                        
                        data_range_numeric = f"{col_letter}2:{col_letter}{len(df) + 1}"
                        
                        # *** ГАРАНТИРОВАННЫЙ РАБОЧИЙ ГРАДИЕНТ (3 ЦВЕТА ARGB) ***
                        rule = ColorScaleRule(
                            start_type='min', start_color='FFFF0000', # Красный (ARGB)
                            mid_type='percentile', mid_value=50, mid_color='FFFFFF00', # Желтый (ARGB)
                            end_type='max', end_color='FF00FF00' # Зеленый (ARGB)
                        )
                        ws.conditional_formatting.add(data_range_numeric, rule)
                        
                        # Правило для максимального значения (Зеленый жирный)
                        max_rule = Rule(type='cellIs', operator='equal', formula=[f'=MAX(${col_letter}$2:${col_letter}${len(df)+1})'])
                        max_rule.dxf = DifferentialStyle(font=Font(bold=True, color='FF008000', size=11))
                        ws.conditional_formatting.add(data_range_numeric, max_rule)
                        
                        # Правило для минимального значения (Красный жирный)
                        min_rule = Rule(type='cellIs', operator='equal', formula=[f'=MIN(${col_letter}$2:${col_letter}${len(df)+1})'])
                        min_rule.dxf = DifferentialStyle(font=Font(bold=True, color='FFFF0000', size=11))
                        ws.conditional_formatting.add(data_range_numeric, min_rule)
                    
                    # Форматирование заголовков
                    ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'{col_letter}1'].font = Font(bold=True, size=11, color='FF000000')

        print("-" * 50)
        print(f"Создан файл отчета: {filepath}, {len(dataframes_dict)} лист(а/ов), {total_rows} строк.")
        print("-" * 50)
    
    except Exception as e:
        print(f"[ОШИБКА EXCEL] Не удалось экспортировать в Excel: {e}")


# --- ТОЧКА ВХОДА ---
if __name__ == "__main__":
    
    # 1. ЗАПУСК АНАЛИТИКИ И ГРАФИКОВ
    generate_all_charts()
    
    # 2. ЗАПУСК PLOTLY (РАСКОММЕНТИРОВАТЬ ДЛЯ ТЕСТИРОВАНИЯ)
    create_plotly_animation()
    
    # 3. ЭКСПОРТ В EXCEL
    query_excel_1 = """
    SELECT 
        t."NICKNAME", 
        r."W" AS total_wins, 
        r."L" AS total_losses, 
        r."W_PCT" AS win_percentage
    FROM 
        ranking r
    INNER JOIN 
        teams t ON r."TEAM_ID" = t."TEAM_ID"
    WHERE 
        r."SEASON_ID" = (SELECT MAX("SEASON_ID") FROM ranking)
    ORDER BY 
        r."W_PCT" DESC
    LIMIT 20;
    """
    
    query_excel_2 = """
    SELECT
        p."PLAYER_NAME",
        SUM(gd."AST") AS total_assists,
        SUM(gd."REB") AS total_rebounds
    FROM
        games_details gd
    INNER JOIN
        players p ON gd."PLAYER_ID" = p."PLAYER_ID"
    GROUP BY
        p."PLAYER_NAME"
    ORDER BY
        total_assists DESC
    LIMIT 20;
    """
    
    df_excel_1 = run_query_to_df(query_excel_1)
    df_excel_2 = run_query_to_df(query_excel_2)

    export_data = {
        'Top_Teams_Stats': df_excel_1,
        'Top_Players_Assists_Reb': df_excel_2
    }
    
    export_to_excel(export_data, 'nba_analytics_report.xlsx')